#!/usr/bin/env python3
"""
ZeitFlow – Telegram Bot für Zeiterfassung

Erweitert um:
- Rollen: Mitarbeiter / Vorarbeiter / Admin
- Fremderfassung für Vorarbeiter/Admins
- Arbeitszeit + Urlaub + Krank
- Korrektur vorhandener Einträge
- Protokollierung: Mitarbeiter + Ersteller + letzte Änderung
- Zwei getrennte Exporte:
  /exportkunde  -> Kundennachweis im SMF-Layout
  /exportintern -> Interne Stunden-/Abwesenheitstabelle im Vadim-Layout
"""

import io
import logging
import os
import re
import sqlite3
from collections import defaultdict
from contextlib import contextmanager
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Callable

from telegram import BotCommand, BotCommandScopeChat, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, ReplyKeyboardRemove, Update
from telegram.ext import (
    Application,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    ConversationHandler,
    MessageHandler,
    filters,
)

try:
    from openpyxl import load_workbook
    XLSX = True
except ImportError:
    XLSX = False


LOG = logging.getLogger("zeitflow")
logging.basicConfig(format="%(asctime)s %(levelname)s %(message)s", level=logging.INFO)

BASE_DIR = Path(__file__).parent
DB_PATH = Path(os.environ.get("ZEITFLOW_DB_PATH", BASE_DIR / "zeitflow.db"))
TEMPLATE_KUNDE_PATH = Path(os.environ.get("ZEITFLOW_TEMPLATE_KUNDE", BASE_DIR / "SMF_2026_KW13.xlsx"))
TEMPLATE_INTERN_PATH = Path(os.environ.get("ZEITFLOW_TEMPLATE_INTERN", BASE_DIR / "Stundentabelle_Vadim.xlsx"))

ABSENCE_PROJECT_NAME = "__ABWESENHEIT__"

MONTH_NAMES_DE = {
    1: "Januar",
    2: "Februar",
    3: "März",
    4: "April",
    5: "Mai",
    6: "Juni",
    7: "Juli",
    8: "August",
    9: "September",
    10: "Oktober",
    11: "November",
    12: "Dezember",
}


def month_sheet_name(month: int, year: int) -> str:
    return f"{MONTH_NAMES_DE[month]} {year}"



@contextmanager
def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def _column_exists(conn: sqlite3.Connection, table: str, column: str) -> bool:
    cols = conn.execute(f"PRAGMA table_info({table})").fetchall()
    return any(r[1] == column for r in cols)


def init_db():
    with db() as c:
        c.executescript(
            """
            CREATE TABLE IF NOT EXISTS users (
                id                INTEGER PRIMARY KEY AUTOINCREMENT,
                tid               INTEGER UNIQUE NOT NULL,
                name              TEXT NOT NULL,
                lang              TEXT DEFAULT 'de',
                is_admin          INTEGER DEFAULT 0,
                active            INTEGER DEFAULT 1,
                created_at        TEXT DEFAULT (datetime('now'))
            );

            CREATE TABLE IF NOT EXISTS projects (
                id                INTEGER PRIMARY KEY AUTOINCREMENT,
                name              TEXT NOT NULL,
                customer          TEXT DEFAULT '',
                cost_center       TEXT DEFAULT '',
                active            INTEGER DEFAULT 1
            );

            CREATE TABLE IF NOT EXISTS entries (
                id                INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id           INTEGER NOT NULL REFERENCES users(id),
                project_id        INTEGER NOT NULL REFERENCES projects(id),
                edate             TEXT NOT NULL,
                stime             TEXT NOT NULL,
                etime             TEXT NOT NULL,
                brk               INTEGER DEFAULT 30,
                hours             REAL NOT NULL,
                notes             TEXT DEFAULT '',
                created_at        TEXT DEFAULT (datetime('now'))
            );

            CREATE INDEX IF NOT EXISTS ix_e_date ON entries(edate);
            CREATE INDEX IF NOT EXISTS ix_e_user ON entries(user_id);
            """
        )

        if not _column_exists(c, "users", "role"):
            c.execute("ALTER TABLE users ADD COLUMN role TEXT DEFAULT 'employee'")
            c.execute("UPDATE users SET role='admin' WHERE is_admin=1")
            c.execute("UPDATE users SET role='employee' WHERE role IS NULL OR role=''")
        if not _column_exists(c, "users", "daily_target_hours"):
            c.execute("ALTER TABLE users ADD COLUMN daily_target_hours REAL DEFAULT 8.0")
        if not _column_exists(c, "users", "vacation_days"):
            c.execute("ALTER TABLE users ADD COLUMN vacation_days REAL DEFAULT 30")

        if not _column_exists(c, "entries", "created_by"):
            c.execute("ALTER TABLE entries ADD COLUMN created_by INTEGER")
            c.execute("UPDATE entries SET created_by=user_id WHERE created_by IS NULL")
        if not _column_exists(c, "entries", "updated_by"):
            c.execute("ALTER TABLE entries ADD COLUMN updated_by INTEGER")
        if not _column_exists(c, "entries", "updated_at"):
            c.execute("ALTER TABLE entries ADD COLUMN updated_at TEXT")
        if not _column_exists(c, "entries", "entry_type"):
            c.execute("ALTER TABLE entries ADD COLUMN entry_type TEXT DEFAULT 'work'")
            c.execute("UPDATE entries SET entry_type='work' WHERE entry_type IS NULL OR entry_type=''")
        if not _column_exists(c, "entries", "billable"):
            c.execute("ALTER TABLE entries ADD COLUMN billable INTEGER DEFAULT 1")
            c.execute("UPDATE entries SET billable=1 WHERE billable IS NULL")
        if not _column_exists(c, "entries", "paid"):
            c.execute("ALTER TABLE entries ADD COLUMN paid INTEGER DEFAULT 1")
            c.execute("UPDATE entries SET paid=1 WHERE paid IS NULL")

        if c.execute("SELECT COUNT(*) FROM projects").fetchone()[0] == 0:
            c.executemany(
                "INSERT INTO projects (name,customer,cost_center) VALUES (?,?,?)",
                [
                    ("Neubau Halle B", "Schneider GmbH", "KST-4010"),
                    ("Sanierung Dach", "Gemeinde Freiburg", "KST-4020"),
                    ("Montage Lüftung", "Klinik Süd", "KST-4030"),
                    ("Heizung Austausch", "Familie Braun", "KST-4040"),
                ],
            )
            LOG.info("Demo-Projekte angelegt")

        absence = c.execute("SELECT id FROM projects WHERE name=?", (ABSENCE_PROJECT_NAME,)).fetchone()
        if not absence:
            c.execute(
                "INSERT INTO projects (name,customer,cost_center,active) VALUES (?,?,?,1)",
                (ABSENCE_PROJECT_NAME, "Intern", "ABW"),
            )

        c.execute("UPDATE users SET role='admin' WHERE is_admin=1 AND (role IS NULL OR role='employee')")


T = {
    "de": {
        "flag": "🇩🇪",
        "label": "Deutsch",
        "welcome": (
            "👋 ZeitFlow\n\n"
            "Hauptmenü:\n"
            "/zeit – Zeit oder Abwesenheit erfassen\n"
            "/heute – Meine heutigen Einträge\n"
            "/woche – Meine Woche anzeigen\n"
            "/projekte – Aktive Projekte ansehen\n"
            "/korrektur – Eintrag ändern oder löschen\n"
            "/sprache – Sprache ändern\n"
            "/hilfe – Übersicht anzeigen"
        ),
        "admin": (
            "\n\n👑 Leitung & Büro:\n"
            "/team – Teamübersicht\n"
            "/stats – Kennzahlen\n"
            "/exportkunde – Kundennachweis exportieren\n"
            "/exportintern – Interne Stundentabelle exportieren\n"
            "/rolle – Mitarbeiterrolle setzen\n"
            "/addprojekt – Projekt anlegen\n"
            "/editprojekt – Projekt bearbeiten\n"
            "/delprojekt – Projekt deaktivieren\n"
            "/deluser – Mitarbeiter deaktivieren"
        ),
        "ask_name": "Vor- und Nachname?",
        "name_ok": "Hallo {name} 👋",
        "pick_target": "Für wen soll der Eintrag angelegt werden?",
        "pick_type": "Was möchtest du erfassen?",
        "type_work": "Arbeit",
        "type_vacation": "Urlaub",
        "type_sick": "Krank",
        "pick_proj": "Welches Projekt?",
        "pick_date": "📅 Datum? (TT.MM.JJJJ / heute)",
        "ask_s": "🕐 Beginn?",
        "ask_e": "🕑 Ende?",
        "ask_b": "☕ Pause (Min)?",
        "ask_n": "📝 Tätigkeit / Bemerkung? (/skip)",
        "pick_abs_mode": "Wie soll die Abwesenheit erfasst werden?",
        "abs_full": "Ganzer Tag",
        "abs_half": "Halber Tag",
        "abs_manual": "Stundenweise",
        "ask_hours": "⏱ Wie viele Stunden?",
        "saved": "✅ Gespeichert",
        "more": "Noch ein Eintrag?",
        "y": "✅ Ja",
        "n": "❌ Fertig",
        "sum_work": "🛠 {p}\n🏢 {c}\n👤 {employee}\n📅 {d}\n🕐 {s}–{e}\n☕ {b}min\n⏱ {h} Std.",
        "sum_abs": "{icon} {etype}\n👤 {employee}\n📅 {d}\n⏱ {h} Std.",
        "created_by": "Erfasst von",
        "updated_by": "Geändert von",
        "no_e": "Keine Einträge.",
        "tot": "Gesamt",
        "hr": "Std.",
        "bad_t": "❌ Format: 7:00",
        "bad_d": "❌ Format: TT.MM.JJJJ",
        "bad_b": "❌ Bitte Minuten als Zahl eingeben.",
        "bad_h": "❌ Bitte Stunden als Zahl eingeben, z. B. 4 oder 7,5.",
        "bad_kw": "❌ Bitte KW/Jahr eingeben, z. B. 13/2026 oder 13.",
        "bad_year": "❌ Bitte ein Jahr eingeben, z. B. 2026.",
        "no_adm": "⚠️ Nur Leitung.",
        "no_foreman": "⚠️ Nur Vorarbeiter oder Admin.",
        "cancel": "Abgebrochen.",
        "no_p": "Keine Projekte.",
        "today": "heute",
        "ap1": "📋 Projektname?",
        "ap2": "🏢 Kunde?",
        "ap3": "🏷 Kostenstelle?",
        "ap_ok": "✅ Projekt:\n📋 {n}\n🏢 {c}\n🏷 {k}",
        "ep_pick": "Bearbeiten?",
        "ep_what": "Was ändern?",
        "ep_val": "Neuer Wert?",
        "ep_ok": "✅ Aktualisiert.",
        "dp_pick": "Deaktivieren?",
        "dp_ask": "⚠️ ‘{n}’ deaktivieren?",
        "dp_ok": "✅ ‘{n}’ deaktiviert.",
        "team_h": "👥 Team:\n",
        "team_r": "{i} {n} – {role} – {h} Std. ({c} Eintr.) {l}\n",
        "role_pick_user": "Wessen Rolle ändern?",
        "role_pick_value": "Welche Rolle soll {name} bekommen?",
        "role_ok": "✅ {name} ist jetzt {role}.",
        "du_pick": "Entfernen?",
        "du_ask": "⚠️ {n} entfernen?",
        "du_ok": "✅ {n} entfernt.",
        "du_none": "Keine.",
        "yes": "✅ Ja",
        "no": "❌ Nein",
        "lang_ok": "✅ Deutsch",
        "export_ok": "📊 Export erstellt.",
        "export_no": "Keine Daten für diesen Export gefunden.",
        "template_missing": "❌ Vorlage nicht gefunden: {name}",
        "corr_pick": "Welchen Eintrag willst du korrigieren?",
        "corr_field": "Was willst du ändern?",
        "corr_value": "Neuen Wert eingeben.",
        "corr_ok": "✅ Eintrag aktualisiert.",
        "corr_deleted": "🗑️ Eintrag gelöscht.",
        "corr_none": "Keine passenden Einträge gefunden.",
        "field_type": "Art",
        "field_date": "Datum",
        "field_start": "Beginn",
        "field_end": "Ende",
        "field_break": "Pause",
        "field_hours": "Stunden",
        "field_notes": "Bemerkung",
        "field_delete": "Löschen",
        "role_employee": "Mitarbeiter",
        "role_foreman": "Vorarbeiter",
        "role_admin": "Admin",
        "self_entry": "🙋 Für mich",
        "self_export": "🙋 Mein Export",
        "skip": "/skip",
        "pick_customer": "Für welchen Kunden soll exportiert werden?",
        "pick_project_export": "Welches Projekt soll exportiert werden?",
        "ask_kw": "Bitte KW/Jahr eingeben, z. B. 13/2026. Nur KW = aktuelles Jahr.",
        "ask_export_year": "Bitte Jahr eingeben, z. B. 2026.",
        "pick_export_user": "Für welchen Mitarbeiter soll exportiert werden?",
        "exportkunde_done": "📄 Kundenexport erstellt.",
        "exportintern_done": "📄 Interner Export erstellt.",
    },
    "ru": {"flag": "🇷🇺", "label": "Русский"},
    "pl": {"flag": "🇵🇱", "label": "Polski"},
    "lv": {"flag": "🇱🇻", "label": "Latviešu"},
}


def t(lang: str, key: str) -> str:
    return T.get(lang, T["de"]).get(key, T["de"].get(key, key))


S_NAME = 0
(
    S_TARGET,
    S_TYPE,
    S_PROJ,
    S_DATE,
    S_START,
    S_END,
    S_BREAK,
    S_NOTES,
    S_MORE,
    S_ABS_MODE,
    S_ABS_HOURS,
) = range(1, 12)
S_AP1, S_AP2, S_AP3 = range(12, 15)
S_EP_PICK, S_EP_FIELD, S_EP_VAL = range(15, 18)
S_DP_PICK, S_DP_CONFIRM = range(18, 20)
S_ROLE_PICK_USER, S_ROLE_PICK_VALUE = range(20, 22)
S_DU_PICK, S_DU_CONFIRM = range(22, 24)
S_CORR_PICK, S_CORR_FIELD, S_CORR_VALUE, S_CORR_DELETE, S_CORR_TYPE = range(24, 29)
S_EK_CUSTOMER, S_EK_PROJECT, S_EK_WEEK = range(29, 32)
S_EI_TARGET, S_EI_YEAR = range(32, 34)


def role_label(role: str, lang: str = "de") -> str:
    return {
        "employee": t(lang, "role_employee"),
        "foreman": t(lang, "role_foreman"),
        "admin": t(lang, "role_admin"),
    }.get(role or "employee", role or "employee")


def type_label(entry_type: str, lang: str = "de") -> str:
    return {
        "work": t(lang, "type_work"),
        "vacation": t(lang, "type_vacation"),
        "sick": t(lang, "type_sick"),
    }.get(entry_type, entry_type)


def type_icon(entry_type: str) -> str:
    return {"work": "🛠", "vacation": "🏖", "sick": "🤒"}.get(entry_type, "•")


def L(ctx):
    return ctx.user_data.get("lang", "de")


def UID(ctx):
    return ctx.user_data.get("uid")


def ROLE(ctx):
    return ctx.user_data.get("role", "employee")


def ADM(ctx):
    return ROLE(ctx) == "admin" or bool(ctx.user_data.get("adm", False))


def FOREMAN(ctx):
    return ROLE(ctx) in {"foreman", "admin"} or ADM(ctx)


def sync_user(ctx, u: dict):
    role = u.get("role") or ("admin" if u.get("is_admin") else "employee")
    ctx.user_data.update(
        {
            "uid": u["id"],
            "lang": u["lang"],
            "adm": bool(u.get("is_admin")),
            "tid": u["tid"],
            "role": role,
        }
    )


def role_menu_text(lang: str, role: str) -> str:
    msg = t(lang, "welcome")
    if role == "foreman":
        msg += (
            "\n\n🦺 Vorarbeiter:\n"
            "/team – Teamübersicht\n"
            "/stats – Kennzahlen\n"
            "/exportkunde – Kundennachweis exportieren\n"
            "/exportintern – Interne Stundentabelle exportieren"
        )
    elif role == "admin":
        msg += (
            "\n\n🦺 Vorarbeiter:\n"
            "/team – Teamübersicht\n"
            "/stats – Kennzahlen\n"
            "/exportkunde – Kundennachweis exportieren\n"
            "/exportintern – Interne Stundentabelle exportieren"
            "\n\n👑 Administration:\n"
            "/rolle – Mitarbeiterrolle setzen\n"
            "/addprojekt – Projekt anlegen\n"
            "/editprojekt – Projekt bearbeiten\n"
            "/delprojekt – Projekt deaktivieren\n"
            "/deluser – Mitarbeiter deaktivieren"
        )
    return msg


def commands_for_role(role: str) -> list[BotCommand]:
    commands = [
        BotCommand("zeit", "Zeit oder Abwesenheit erfassen"),
        BotCommand("heute", "Heutige Einträge anzeigen"),
        BotCommand("woche", "Wochensumme und Einträge anzeigen"),
        BotCommand("projekte", "Aktive Projekte anzeigen"),
        BotCommand("korrektur", "Eintrag ändern oder löschen"),
        BotCommand("sprache", "Sprache ändern"),
        BotCommand("hilfe", "Befehlsübersicht anzeigen"),
    ]
    if role in {"foreman", "admin"}:
        commands.extend(
            [
                BotCommand("team", "Teamübersicht anzeigen"),
                BotCommand("stats", "Kennzahlen anzeigen"),
                BotCommand("exportkunde", "Kundennachweis exportieren"),
                BotCommand("exportintern", "Interne Stundentabelle exportieren"),
            ]
        )
    if role == "admin":
        commands.extend(
            [
                BotCommand("rolle", "Mitarbeiterrolle setzen"),
                BotCommand("addprojekt", "Projekt anlegen"),
                BotCommand("editprojekt", "Projekt bearbeiten"),
                BotCommand("delprojekt", "Projekt deaktivieren"),
                BotCommand("deluser", "Mitarbeiter deaktivieren"),
            ]
        )
    return commands


async def apply_user_menu(bot, tid: int, role: str) -> None:
    await bot.set_my_commands(commands_for_role(role), scope=BotCommandScopeChat(chat_id=tid))


def detect_template_year(wb) -> int:
    for title in wb.sheetnames:
        m = re.fullmatch(r"(\d{4})", str(title))
        if m:
            return int(m.group(1))
    for title in wb.sheetnames:
        m = re.search(r"(20\d{2})", str(title))
        if m:
            return int(m.group(1))
    return date.today().year


def normalise_internal_template_year(wb, target_year: int) -> tuple[str, dict[int, str]]:
    template_year = detect_template_year(wb)
    year_sheet_old = next((s for s in wb.sheetnames if re.fullmatch(r"\d{4}", str(s))), wb.sheetnames[0])
    year_sheet_new = str(target_year)
    if year_sheet_old != year_sheet_new:
        wb[year_sheet_old].title = year_sheet_new

    month_mapping: dict[str, str] = {}
    month_sheets: dict[int, str] = {}
    for month, month_label in MONTH_NAMES_DE.items():
        old_title = next((s for s in wb.sheetnames if s == f"{month_label} {template_year}"), None)
        if old_title is None:
            old_title = next((s for s in wb.sheetnames if str(s).startswith(f"{month_label} ")), None)
        if old_title is None:
            continue
        new_title = month_sheet_name(month, target_year)
        if old_title != new_title:
            wb[old_title].title = new_title
        month_mapping[old_title] = new_title
        month_sheets[month] = new_title

    replacements = {year_sheet_old: year_sheet_new, **month_mapping}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str):
                    continue
                new_value = value
                for old_name, new_name in replacements.items():
                    new_value = new_value.replace(f"'{old_name}'!", f"'{new_name}'!")
                    new_value = new_value.replace(f"{old_name}!", f"{new_name}!")
                if cell.coordinate == "A2" and "Jahresliste" in new_value:
                    new_value = f"Jahresliste {target_year}"
                cell.value = new_value

    for defined_name in wb.defined_names.values():
        attr = getattr(defined_name, "attr_text", None)
        if not attr:
            continue
        new_attr = attr
        for old_name, new_name in replacements.items():
            new_attr = new_attr.replace(f"'{old_name}'!", f"'{new_name}'!")
            new_attr = new_attr.replace(f"{old_name}!", f"{new_name}!")
        defined_name.attr_text = new_attr

    return year_sheet_new, month_sheets


def parse_time(s: str) -> str | None:
    s = s.strip().replace(".", ":").replace(",", ":")
    for fmt in ("%H:%M", "%H"):
        try:
            d = datetime.strptime(s, fmt)
            return f"{d.hour:02d}:{d.minute:02d}"
        except ValueError:
            pass
    return None


def parse_date(s: str) -> str | None:
    s = s.strip().lower()
    if s in ("heute", "today", "сегодня", "dzisiaj", "šodien"):
        return date.today().isoformat()
    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def parse_hours(s: str) -> float | None:
    raw = s.strip().lower().replace("std", "").replace("stunden", "").replace("h", "").replace(",", ".").strip()
    try:
        value = float(raw)
    except ValueError:
        return None
    if value < 0 or value > 24:
        return None
    return round(value, 2)


def parse_kw_year(s: str) -> tuple[int, int] | None:
    raw = s.strip().lower().replace("kw", "").replace(" ", "")
    if "/" in raw:
        left, right = raw.split("/", 1)
        if left.isdigit() and right.isdigit():
            kw, year = int(left), int(right)
        else:
            return None
    else:
        if not raw.isdigit():
            return None
        kw, year = int(raw), date.today().year
    if not (1 <= kw <= 53 and 2000 <= year <= 2100):
        return None
    try:
        date.fromisocalendar(year, kw, 1)
    except ValueError:
        return None
    return kw, year


def fde(d: str) -> str:
    try:
        return datetime.strptime(d, "%Y-%m-%d").strftime("%d.%m.%Y")
    except Exception:
        return d


def calc_hours(stime: str, etime: str, brk: int) -> float:
    sh, sm = map(int, stime.split(":"))
    eh, em = map(int, etime.split(":"))
    mins = (eh * 60 + em) - (sh * 60 + sm) - brk
    return round(max(0, mins) / 60, 2)


def end_time_from_hours(hours: float) -> str:
    mins = int(round(hours * 60))
    h, m = divmod(mins, 60)
    h = max(0, min(23, h))
    return f"{h:02d}:{m:02d}"


def excel_time_from_hhmm(value: str):
    if not value:
        return None
    try:
        hh, mm = map(int, value.split(":"))
        return time(hh, mm)
    except Exception:
        return None


def excel_time_from_break_minutes(minutes: int):
    minutes = max(0, int(minutes or 0))
    hh, mm = divmod(minutes, 60)
    hh = max(0, min(23, hh))
    return time(hh, mm)


def get_or_create_user(tid: int, name: str, lang: str = "de") -> dict:
    with db() as c:
        u = c.execute("SELECT * FROM users WHERE tid=?", (tid,)).fetchone()
        if u:
            return dict(u)
        has_admin = c.execute("SELECT COUNT(*) FROM users WHERE role='admin' OR is_admin=1").fetchone()[0]
        role = "employee" if has_admin else "admin"
        is_admin = 1 if role == "admin" else 0
        c.execute(
            "INSERT INTO users (tid,name,lang,is_admin,role) VALUES (?,?,?,?,?)",
            (tid, name, lang, is_admin, role),
        )
        return dict(c.execute("SELECT * FROM users WHERE tid=?", (tid,)).fetchone())


def get_user(tid: int) -> dict | None:
    with db() as c:
        u = c.execute("SELECT * FROM users WHERE tid=? AND active=1", (tid,)).fetchone()
        return dict(u) if u else None


def get_user_by_id(user_id: int) -> dict | None:
    with db() as c:
        u = c.execute("SELECT * FROM users WHERE id=? AND active=1", (user_id,)).fetchone()
        return dict(u) if u else None


def get_users(active_only: bool = True) -> list[dict]:
    sql = "SELECT * FROM users"
    if active_only:
        sql += " WHERE active=1"
    sql += " ORDER BY name"
    with db() as c:
        return [dict(r) for r in c.execute(sql).fetchall()]


def get_absence_project_id() -> int:
    with db() as c:
        row = c.execute("SELECT id FROM projects WHERE name=?", (ABSENCE_PROJECT_NAME,)).fetchone()
        if row:
            return int(row["id"])
    raise RuntimeError("Abwesenheitsprojekt fehlt")


def get_projects(include_system: bool = False) -> list[dict]:
    with db() as c:
        if include_system:
            rows = c.execute("SELECT * FROM projects WHERE active=1 ORDER BY customer, name").fetchall()
        else:
            rows = c.execute(
                "SELECT * FROM projects WHERE active=1 AND name<>? ORDER BY customer, name",
                (ABSENCE_PROJECT_NAME,),
            ).fetchall()
        return [dict(r) for r in rows]


def get_project(project_id: int) -> dict | None:
    with db() as c:
        row = c.execute("SELECT * FROM projects WHERE id=?", (project_id,)).fetchone()
        return dict(row) if row else None


def project_short_name(project_name: str) -> str:
    name = (project_name or "").strip()
    if not name:
        return ""
    if len(name) <= 14:
        return name
    parts = [p for p in name.replace("-", " ").split() if p]
    if len(parts) >= 2:
        short = "/".join(p[:6] for p in parts[:3])
        return short[:30]
    return name[:30]


def current_user_row(ctx) -> dict | None:
    return get_user_by_id(UID(ctx)) if UID(ctx) else None
PROJECT_PAGE_SIZE = 20

def build_project_keyboard(ctx, page: int = 1):
    """Returns keyboard for project selection with pagination."""
    ps = get_projects()
    total = len(ps)
    max_page = max(1, (total + PROJECT_PAGE_SIZE - 1) // PROJECT_PAGE_SIZE)
    page = max(1, min(page, max_page))

    start = (page - 1) * PROJECT_PAGE_SIZE
    chunk = ps[start:start + PROJECT_PAGE_SIZE]

    kb = [[InlineKeyboardButton(f"{p['name']} – {p['customer']} ({p['cost_center']})", callback_data=f"p_{p['id']}")]
          for p in chunk]

    nav = []
    if page > 1:
        nav.append(InlineKeyboardButton("←", callback_data=f"proj_page_{page-1}"))
    if page < max_page:
        nav.append(InlineKeyboardButton("→", callback_data=f"proj_page_{page+1}"))
    if nav:
        kb.append(nav)

    ctx.user_data["proj_page"] = page
    ctx.user_data["proj_max_page"] = max_page
    return InlineKeyboardMarkup(kb), page, max_page, total



def save_entry(
    *,
    user_id: int,
    created_by: int,
    project_id: int,
    edate: str,
    entry_type: str,
    notes: str = "",
    stime: str = "00:00",
    etime: str = "00:00",
    brk: int = 0,
    hours: float | None = None,
) -> float:
    if entry_type == "work":
        hours = calc_hours(stime, etime, brk)
        billable = 1
        paid = 1
    else:
        hours = round(float(hours or 0), 2)
        stime = "00:00"
        etime = end_time_from_hours(hours)
        brk = 0
        billable = 0
        paid = 1

    with db() as c:
        c.execute(
            """
            INSERT INTO entries (
                user_id, created_by, project_id, edate, stime, etime, brk, hours, notes,
                entry_type, billable, paid
            )
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (user_id, created_by, project_id, edate, stime, etime, brk, hours, notes, entry_type, billable, paid),
        )
    return hours


def get_entry(entry_id: int) -> dict | None:
    with db() as c:
        row = c.execute(
            """
            SELECT e.*, u.name AS employee, p.name AS project, p.customer, p.cost_center,
                   cu.name AS creator_name, uu.name AS updater_name
            FROM entries e
            JOIN users u ON e.user_id=u.id
            JOIN projects p ON e.project_id=p.id
            LEFT JOIN users cu ON cu.id=e.created_by
            LEFT JOIN users uu ON uu.id=e.updated_by
            WHERE e.id=?
            """,
            (entry_id,),
        ).fetchone()
        return dict(row) if row else None


def get_entries(
    *,
    user_id: int | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    project_id: int | None = None,
    entry_type: str | None = None,
    limit: int | None = None,
    editable_by: dict | None = None,
) -> list[dict]:
    conds = ["1=1"]
    params: list = []
    if user_id:
        conds.append("e.user_id=?")
        params.append(user_id)
    if date_from:
        conds.append("e.edate>=?")
        params.append(date_from)
    if date_to:
        conds.append("e.edate<=?")
        params.append(date_to)
    if project_id:
        conds.append("e.project_id=?")
        params.append(project_id)
    if entry_type:
        conds.append("e.entry_type=?")
        params.append(entry_type)
    if editable_by and editable_by.get("role") not in {"foreman", "admin"}:
        conds.append("e.user_id=?")
        params.append(editable_by["id"])
    sql = f"""
        SELECT e.*, u.name AS employee, p.name AS project, p.customer, p.cost_center,
               cu.name AS creator_name, uu.name AS updater_name
        FROM entries e
        JOIN users u ON e.user_id=u.id
        JOIN projects p ON e.project_id=p.id
        LEFT JOIN users cu ON cu.id=e.created_by
        LEFT JOIN users uu ON uu.id=e.updated_by
        WHERE {' AND '.join(conds)}
        ORDER BY e.edate DESC, u.name, e.stime, e.id DESC
    """
    if limit:
        sql += f" LIMIT {int(limit)}"
    with db() as c:
        return [dict(r) for r in c.execute(sql, params).fetchall()]


def update_entry(entry_id: int, actor_id: int, **fields):
    current = get_entry(entry_id)
    if not current:
        return

    for key, value in fields.items():
        current[key] = value

    if current["entry_type"] == "work":
        current["billable"] = 1
        current["paid"] = 1
        current["hours"] = calc_hours(current["stime"], current["etime"], int(current["brk"]))
    else:
        current["billable"] = 0
        current["paid"] = 1
        if "hours" in fields:
            current["hours"] = round(float(current["hours"]), 2)
        current["brk"] = 0
        current["stime"] = "00:00"
        current["etime"] = end_time_from_hours(float(current["hours"]))

    set_cols = [f"{col}=?" for col in fields]
    values = [current[col] for col in fields]
    set_cols.extend(["hours=?", "billable=?", "paid=?", "stime=?", "etime=?", "brk=?", "updated_by=?", "updated_at=?"])
    values.extend(
        [
            current["hours"],
            current["billable"],
            current["paid"],
            current["stime"],
            current["etime"],
            current["brk"],
            actor_id,
            datetime.utcnow().isoformat(timespec="seconds"),
            entry_id,
        ]
    )
    sql = f"UPDATE entries SET {', '.join(set_cols)} WHERE id=?"
    with db() as c:
        c.execute(sql, values)


def delete_entry(entry_id: int):
    with db() as c:
        c.execute("DELETE FROM entries WHERE id=?", (entry_id,))


def entry_brief(e: dict, lang: str = "de") -> str:
    if e["entry_type"] == "work":
        return (
            f"{type_icon('work')} {fde(e['edate'])} | {e['employee']} | {e['project']} | "
            f"{e['stime']}-{e['etime']} | {e['hours']:.1f} {t(lang, 'hr')}"
        )
    return (
        f"{type_icon(e['entry_type'])} {fde(e['edate'])} | {e['employee']} | "
        f"{type_label(e['entry_type'], lang)} | {e['hours']:.1f} {t(lang, 'hr')}"
    )


async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    u = get_user(update.effective_user.id)
    if u:
        sync_user(ctx, u)
        await apply_user_menu(ctx.bot, u["tid"], ctx.user_data.get("role", "employee"))
        await update.message.reply_text(role_menu_text(u["lang"], ctx.user_data.get("role", "employee")))
    else:
        kb = [[InlineKeyboardButton(f"{v['flag']} {v['label']}", callback_data=f"rl_{k}")] for k, v in T.items()]
        await update.message.reply_text("🌍 Sprache?", reply_markup=InlineKeyboardMarkup(kb))


async def reg_lang(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    ctx.user_data["reg_lang"] = q.data[3:]
    await q.edit_message_text(t(ctx.user_data["reg_lang"], "ask_name"))
    return S_NAME


async def reg_name(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    name = update.message.text.strip()
    lang = ctx.user_data.get("reg_lang", "de")
    u = get_or_create_user(update.effective_user.id, name, lang)
    sync_user(ctx, u)
    await apply_user_menu(ctx.bot, u["tid"], ctx.user_data.get("role", "employee"))
    hint = "\n\n👑 Du bist Leitung!" if u["role"] == "admin" else ("\n\n🦺 Du bist Vorarbeiter!" if u["role"] == "foreman" else "")
    await update.message.reply_text(t(lang, "name_ok").format(name=name) + hint)
    await update.message.reply_text(role_menu_text(lang, ctx.user_data.get("role", "employee")))
    return ConversationHandler.END


async def cancel(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.message:
        await update.message.reply_text(t(L(ctx), "cancel"), reply_markup=ReplyKeyboardRemove())
    elif update.callback_query:
        await update.callback_query.answer()
        await update.callback_query.edit_message_text(t(L(ctx), "cancel"))
    return ConversationHandler.END


async def _prompt_type(reply_func: Callable, ctx):
    kb = InlineKeyboardMarkup(
        [
            [InlineKeyboardButton(f"🛠 {t(L(ctx), 'type_work')}", callback_data="ty_work")],
            [InlineKeyboardButton(f"🏖 {t(L(ctx), 'type_vacation')}", callback_data="ty_vacation")],
            [InlineKeyboardButton(f"🤒 {t(L(ctx), 'type_sick')}", callback_data="ty_sick")],
        ]
    )
    await reply_func(t(L(ctx), "pick_type"), reply_markup=kb)
    return S_TYPE




CUSTOMER_PAGE_SIZE = 20
PROJECT_EXPORT_PAGE_SIZE = 20
USER_PAGE_SIZE = 20

def build_customer_keyboard(customers: list[str], page: int = 1):
    total = len(customers)
    max_page = max(1, (total + CUSTOMER_PAGE_SIZE - 1) // CUSTOMER_PAGE_SIZE)
    page = max(1, min(page, max_page))
    start = (page - 1) * CUSTOMER_PAGE_SIZE
    chunk = customers[start:start + CUSTOMER_PAGE_SIZE]

    kb = [[InlineKeyboardButton(customer, callback_data=f"ec_{idx}")] for idx, customer in enumerate(chunk, start=start)]

    nav = []
    if page > 1:
        nav.append(InlineKeyboardButton("←", callback_data=f"ecp_{page-1}"))
    if page < max_page:
        nav.append(InlineKeyboardButton("→", callback_data=f"ecp_{page+1}"))
    if nav:
        kb.append(nav)
    return InlineKeyboardMarkup(kb), page, max_page, total


def build_export_project_keyboard(projects: list[dict], page: int = 1):
    total = len(projects)
    max_page = max(1, (total + PROJECT_EXPORT_PAGE_SIZE - 1) // PROJECT_EXPORT_PAGE_SIZE)
    page = max(1, min(page, max_page))
    start = (page - 1) * PROJECT_EXPORT_PAGE_SIZE
    chunk = projects[start:start + PROJECT_EXPORT_PAGE_SIZE]

    kb = [[InlineKeyboardButton(p["name"], callback_data=f"epj_{p['id']}")] for p in chunk]

    nav = []
    if page > 1:
        nav.append(InlineKeyboardButton("←", callback_data=f"epj_page_{page-1}"))
    if page < max_page:
        nav.append(InlineKeyboardButton("→", callback_data=f"epj_page_{page+1}"))
    if nav:
        kb.append(nav)
    return InlineKeyboardMarkup(kb), page, max_page, total


def build_export_user_keyboard(users: list[dict], page: int = 1):
    total = len(users)
    max_page = max(1, (total + USER_PAGE_SIZE - 1) // USER_PAGE_SIZE)
    page = max(1, min(page, max_page))
    start = (page - 1) * USER_PAGE_SIZE
    chunk = users[start:start + USER_PAGE_SIZE]

    kb = [[InlineKeyboardButton(f"👤 {u['name']}", callback_data=f"iu_{u['id']}")] for u in chunk]

    nav = []
    if page > 1:
        nav.append(InlineKeyboardButton("←", callback_data=f"iu_page_{page-1}"))
    if page < max_page:
        nav.append(InlineKeyboardButton("→", callback_data=f"iu_page_{page+1}"))
    if nav:
        kb.append(nav)
    return InlineKeyboardMarkup(kb), page, max_page, total
async def _prompt_project(reply_func: Callable, ctx, page: int | None = None):
    ps = get_projects()
    if not ps:
        await reply_func(t(L(ctx), "no_p"))
        return ConversationHandler.END

    if page is None:
        page = ctx.user_data.get("proj_page", 1)

    kb, page, max_page, total = build_project_keyboard(ctx, page)
    await reply_func(f"{t(L(ctx), 'pick_proj')} ({page}/{max_page}, {total} Projekte)", reply_markup=kb)
    return S_PROJ

async def proj_page_cb(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Navigate project list pages."""
    q = update.callback_query
    await q.answer()
    try:
        page = int(q.data.split("_")[-1])
    except Exception:
        page = 1
    kb, page, max_page, total = build_project_keyboard(ctx, page)
    await q.edit_message_text(f"{t(L(ctx), 'pick_proj')} ({page}/{max_page}, {total} Projekte)", reply_markup=kb)
    return S_PROJ




async def cmd_zeit(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not UID(ctx):
        await update.message.reply_text("→ /start")
        return ConversationHandler.END

    for key in ["target_uid", "target_name", "entry_type", "pid", "pn", "pc", "pk", "ed", "st", "et", "bk", "abs_hours"]:
        ctx.user_data.pop(key, None)

    current = current_user_row(ctx)
    if FOREMAN(ctx):
        users = get_users()
        kb = [[InlineKeyboardButton(t(L(ctx), "self_entry"), callback_data="tu_self")]]
        kb.extend([[InlineKeyboardButton(f"👤 {u['name']}", callback_data=f"tu_{u['id']}")] for u in users if u["id"] != UID(ctx)])
        await update.message.reply_text(t(L(ctx), "pick_target"), reply_markup=InlineKeyboardMarkup(kb))
        return S_TARGET

    ctx.user_data["target_uid"] = UID(ctx)
    ctx.user_data["target_name"] = current["name"] if current else "?"
    return await _prompt_type(update.message.reply_text, ctx)


async def z_target(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    payload = q.data[3:]
    if payload == "self":
        target = current_user_row(ctx)
    else:
        target = get_user_by_id(int(payload))
    if not target:
        await q.edit_message_text("Mitarbeiter nicht gefunden.")
        return ConversationHandler.END
    ctx.user_data["target_uid"] = target["id"]
    ctx.user_data["target_name"] = target["name"]
    return await _prompt_type(q.edit_message_text, ctx)


async def z_type(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    entry_type = q.data[3:]
    ctx.user_data["entry_type"] = entry_type
    kb = [[InlineKeyboardButton(f"📅 {t(L(ctx), 'today').capitalize()}", callback_data="dt")]]
    if entry_type == "work":
        return await _prompt_project(q.edit_message_text, ctx)
    await q.edit_message_text(t(L(ctx), "pick_date"), reply_markup=InlineKeyboardMarkup(kb))
    return S_DATE


async def z_proj(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    pid = int(q.data[2:])
    p = get_project(pid)
    if not p:
        await q.edit_message_text(t(L(ctx), "no_p"))
        return ConversationHandler.END
    ctx.user_data.update({"pid": pid, "pn": p["name"], "pc": p["customer"], "pk": p["cost_center"]})
    kb = [[InlineKeyboardButton(f"📅 {t(L(ctx), 'today').capitalize()}", callback_data="dt")]]
    await q.edit_message_text(t(L(ctx), "pick_date"), reply_markup=InlineKeyboardMarkup(kb))
    return S_DATE


async def z_date_btn(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    ctx.user_data["ed"] = date.today().isoformat()
    if ctx.user_data.get("entry_type") == "work":
        await q.edit_message_text(t(L(ctx), "ask_s"))
        return S_START
    kb = InlineKeyboardMarkup(
        [
            [InlineKeyboardButton(f"🟢 {t(L(ctx), 'abs_full')}", callback_data="am_full")],
            [InlineKeyboardButton(f"🟡 {t(L(ctx), 'abs_half')}", callback_data="am_half")],
            [InlineKeyboardButton(f"🔢 {t(L(ctx), 'abs_manual')}", callback_data="am_manual")],
        ]
    )
    await q.edit_message_text(t(L(ctx), "pick_abs_mode"), reply_markup=kb)
    return S_ABS_MODE


async def z_date_txt(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    d = parse_date(update.message.text)
    if not d:
        await update.message.reply_text(t(L(ctx), "bad_d"))
        return S_DATE
    ctx.user_data["ed"] = d
    if ctx.user_data.get("entry_type") == "work":
        await update.message.reply_text(t(L(ctx), "ask_s"))
        return S_START
    kb = InlineKeyboardMarkup(
        [
            [InlineKeyboardButton(f"🟢 {t(L(ctx), 'abs_full')}", callback_data="am_full")],
            [InlineKeyboardButton(f"🟡 {t(L(ctx), 'abs_half')}", callback_data="am_half")],
            [InlineKeyboardButton(f"🔢 {t(L(ctx), 'abs_manual')}", callback_data="am_manual")],
        ]
    )
    await update.message.reply_text(t(L(ctx), "pick_abs_mode"), reply_markup=kb)
    return S_ABS_MODE


async def z_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    tm = parse_time(update.message.text)
    if not tm:
        await update.message.reply_text(t(L(ctx), "bad_t"))
        return S_START
    ctx.user_data["st"] = tm
    await update.message.reply_text(t(L(ctx), "ask_e"))
    return S_END


async def z_end(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    tm = parse_time(update.message.text)
    if not tm:
        await update.message.reply_text(t(L(ctx), "bad_t"))
        return S_END
    ctx.user_data["et"] = tm
    kb = ReplyKeyboardMarkup([["0", "15", "30", "45", "60"]], one_time_keyboard=True, resize_keyboard=True)
    await update.message.reply_text(t(L(ctx), "ask_b"), reply_markup=kb)
    return S_BREAK


async def z_break(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    try:
        minutes = int(update.message.text.strip())
    except ValueError:
        await update.message.reply_text(t(L(ctx), "bad_b"))
        return S_BREAK
    ctx.user_data["bk"] = minutes
    await update.message.reply_text(t(L(ctx), "ask_n"), reply_markup=ReplyKeyboardRemove())
    return S_NOTES


async def z_abs_mode(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    mode = q.data[3:]
    target = get_user_by_id(ctx.user_data["target_uid"])
    daily_target = float(target.get("daily_target_hours") or 8.0) if target else 8.0
    if mode == "full":
        ctx.user_data["abs_hours"] = round(daily_target, 2)
        await q.edit_message_text(t(L(ctx), "ask_n"))
        return S_NOTES
    if mode == "half":
        ctx.user_data["abs_hours"] = round(daily_target / 2, 2)
        await q.edit_message_text(t(L(ctx), "ask_n"))
        return S_NOTES
    await q.edit_message_text(t(L(ctx), "ask_hours"))
    return S_ABS_HOURS


async def z_abs_hours(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    hours = parse_hours(update.message.text)
    if hours is None:
        await update.message.reply_text(t(L(ctx), "bad_h"))
        return S_ABS_HOURS
    ctx.user_data["abs_hours"] = hours
    await update.message.reply_text(t(L(ctx), "ask_n"))
    return S_NOTES


async def z_notes(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ud = ctx.user_data
    lang = L(ctx)
    notes = "" if update.message.text.strip() == t(lang, "skip") or update.message.text.strip() == "/skip" else update.message.text.strip()
    target = get_user_by_id(ud["target_uid"])
    if not target:
        await update.message.reply_text("Mitarbeiter nicht gefunden.")
        return ConversationHandler.END

    if ud["entry_type"] == "work":
        hours = save_entry(
            user_id=ud["target_uid"],
            created_by=UID(ctx),
            project_id=ud["pid"],
            edate=ud["ed"],
            entry_type="work",
            stime=ud["st"],
            etime=ud["et"],
            brk=ud["bk"],
            notes=notes,
        )
        txt = t(lang, "saved") + "\n\n" + t(lang, "sum_work").format(
            p=ud["pn"],
            c=ud["pc"],
            employee=target["name"],
            d=fde(ud["ed"]),
            s=ud["st"],
            e=ud["et"],
            b=ud["bk"],
            h=f"{hours:.1f}",
        )
    else:
        hours = save_entry(
            user_id=ud["target_uid"],
            created_by=UID(ctx),
            project_id=get_absence_project_id(),
            edate=ud["ed"],
            entry_type=ud["entry_type"],
            hours=float(ud["abs_hours"]),
            notes=notes,
        )
        txt = t(lang, "saved") + "\n\n" + t(lang, "sum_abs").format(
            icon=type_icon(ud["entry_type"]),
            etype=type_label(ud["entry_type"], lang),
            employee=target["name"],
            d=fde(ud["ed"]),
            h=f"{hours:.1f}",
        )

    if notes:
        txt += f"\n📝 {notes}"
    txt += f"\n\n{t(lang, 'created_by')}: {current_user_row(ctx)['name']}"
    kb = InlineKeyboardMarkup(
        [[InlineKeyboardButton(t(lang, "y"), callback_data="my"), InlineKeyboardButton(t(lang, "n"), callback_data="mn")]]
    )
    await update.message.reply_text(txt + "\n\n" + t(lang, "more"), reply_markup=kb)
    return S_MORE


async def z_more(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    if q.data == "my":
        for key in ["entry_type", "pid", "pn", "pc", "pk", "ed", "st", "et", "bk", "abs_hours"]:
            ctx.user_data.pop(key, None)
        return await _prompt_type(q.edit_message_text, ctx)
    await q.edit_message_text("👍")
    return ConversationHandler.END


async def cmd_heute(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not UID(ctx):
        await update.message.reply_text("→ /start")
        return
    lang = L(ctx)
    today = date.today().isoformat()
    es = get_entries(user_id=UID(ctx), date_from=today, date_to=today)
    if not es:
        await update.message.reply_text(t(lang, "no_e"))
        return
    txt = f"📊 {fde(today)}\n"
    total = 0.0
    for idx, e in enumerate(es, 1):
        if e["entry_type"] == "work":
            txt += (
                f"\n{idx}. 🛠 {e['project']} ({e['customer']})\n"
                f"   {e['stime']}–{e['etime']} ☕{e['brk']}min ⏱{e['hours']:.1f}{t(lang, 'hr')}"
            )
        else:
            txt += f"\n{idx}. {type_icon(e['entry_type'])} {type_label(e['entry_type'], lang)} ⏱{e['hours']:.1f}{t(lang, 'hr')}"
        if e["notes"]:
            txt += f"\n   📝 {e['notes']}"
        total += e["hours"]
    txt += f"\n\n━━━━━━━━━━\n{t(lang, 'tot')}: {total:.1f} {t(lang, 'hr')}"
    await update.message.reply_text(txt)


async def cmd_woche(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not UID(ctx):
        await update.message.reply_text("→ /start")
        return
    lang = L(ctx)
    ws = (date.today() - timedelta(days=date.today().weekday())).isoformat()
    es = get_entries(user_id=UID(ctx), date_from=ws)
    if not es:
        await update.message.reply_text(t(lang, "no_e"))
        return
    txt, cur, total = "📊 Woche\n", "", 0.0
    for e in es:
        if e["edate"] != cur:
            cur = e["edate"]
            txt += f"\n📅 {fde(cur)}\n"
        if e["entry_type"] == "work":
            txt += f"  • 🛠 {e['project']} | {e['stime']}-{e['etime']} | {e['hours']:.1f}{t(lang, 'hr')}\n"
        else:
            txt += f"  • {type_icon(e['entry_type'])} {type_label(e['entry_type'], lang)} | {e['hours']:.1f}{t(lang, 'hr')}\n"
        total += e["hours"]
    txt += f"\n━━━━━━━━━━\n{t(lang, 'tot')}: {total:.1f} {t(lang, 'hr')}"
    await update.message.reply_text(txt)


async def cmd_projekte(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ps = get_projects()
    if not ps:
        await update.message.reply_text(t(L(ctx), "no_p"))
        return
    txt = "📋 Projekte\n"
    for p in ps:
        txt += f"\n📋 {p['name']}\n   🏢 {p['customer']}\n   🏷 {p['cost_center']}\n"
    await update.message.reply_text(txt)


async def cmd_sprache(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    kb = [[InlineKeyboardButton(f"{v['flag']} {v['label']}", callback_data=f"sl_{k}")] for k, v in T.items()]
    await update.message.reply_text("🌍", reply_markup=InlineKeyboardMarkup(kb))


async def sprache_cb(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    lang = q.data[3:]
    ctx.user_data["lang"] = lang
    u = get_user(q.from_user.id)
    if u:
        with db() as c:
            c.execute("UPDATE users SET lang=? WHERE id=?", (lang, u["id"]))
    await q.edit_message_text(t(lang, "lang_ok"))


async def cmd_hilfe(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(role_menu_text(L(ctx), ROLE(ctx)))


async def cmd_team(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not FOREMAN(ctx):
        await update.message.reply_text(t(L(ctx), "no_foreman"))
        return
    with db() as c:
        users = [dict(r) for r in c.execute("SELECT * FROM users WHERE active=1 ORDER BY name").fetchall()]
        txt = t(L(ctx), "team_h")
        for u in users:
            s = c.execute(
                "SELECT COUNT(*) AS c, COALESCE(SUM(hours),0) AS h FROM entries WHERE user_id=?",
                (u["id"],),
            ).fetchone()
            txt += t(L(ctx), "team_r").format(
                i="👑" if u["role"] == "admin" else ("🦺" if u["role"] == "foreman" else "👤"),
                n=u["name"],
                role=role_label(u["role"], L(ctx)),
                h=f"{s['h']:.1f}",
                c=s["c"],
                l=T.get(u["lang"], T["de"])["flag"],
            )
    await update.message.reply_text(txt)


async def cmd_stats(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not FOREMAN(ctx):
        await update.message.reply_text(t(L(ctx), "no_foreman"))
        return
    today = date.today().isoformat()
    ws = (date.today() - timedelta(days=date.today().weekday())).isoformat()
    ms = date.today().replace(day=1).isoformat()
    with db() as c:
        td = c.execute(
            "SELECT COALESCE(SUM(hours),0) AS h, COUNT(*) AS c FROM entries WHERE edate=?",
            (today,),
        ).fetchone()
        wk = c.execute(
            "SELECT COALESCE(SUM(hours),0) AS h, COUNT(*) AS c FROM entries WHERE edate>=?",
            (ws,),
        ).fetchone()
        mo = c.execute(
            "SELECT COALESCE(SUM(hours),0) AS h, COUNT(*) AS c FROM entries WHERE edate>=?",
            (ms,),
        ).fetchone()
        absm = c.execute(
            """
            SELECT entry_type, COALESCE(SUM(hours),0) AS h
            FROM entries
            WHERE edate>=? AND entry_type IN ('vacation','sick')
            GROUP BY entry_type
            """,
            (ms,),
        ).fetchall()
    abs_map = {r["entry_type"]: r["h"] for r in absm}
    txt = (
        f"📊 Dashboard\n\n"
        f"Heute: {td['h']:.1f} Std. ({td['c']} Einträge)\n"
        f"Woche: {wk['h']:.1f} Std. ({wk['c']} Einträge)\n"
        f"Monat: {mo['h']:.1f} Std. ({mo['c']} Einträge)\n"
        f"🏖 Urlaub (Monat): {abs_map.get('vacation', 0):.1f} Std.\n"
        f"🤒 Krank (Monat): {abs_map.get('sick', 0):.1f} Std."
    )
    await update.message.reply_text(txt)


async def cmd_addprojekt(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not ADM(ctx):
        await update.message.reply_text(t(L(ctx), "no_adm"))
        return ConversationHandler.END
    await update.message.reply_text(t(L(ctx), "ap1"))
    return S_AP1


async def ap1(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data["_apn"] = update.message.text.strip()
    await update.message.reply_text(t(L(ctx), "ap2"))
    return S_AP2


async def ap2(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data["_apc"] = update.message.text.strip()
    await update.message.reply_text(t(L(ctx), "ap3"))
    return S_AP3


async def ap3(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ud = ctx.user_data
    k = update.message.text.strip()
    with db() as c:
        c.execute("INSERT INTO projects (name,customer,cost_center) VALUES (?,?,?)", (ud["_apn"], ud["_apc"], k))
    await update.message.reply_text(t(L(ctx), "ap_ok").format(n=ud["_apn"], c=ud["_apc"], k=k))
    return ConversationHandler.END


async def cmd_editprojekt(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not ADM(ctx):
        await update.message.reply_text(t(L(ctx), "no_adm"))
        return ConversationHandler.END
    ps = get_projects()
    if not ps:
        await update.message.reply_text(t(L(ctx), "no_p"))
        return ConversationHandler.END
    kb = [[InlineKeyboardButton(f"{p['name']} – {p['customer']}", callback_data=f"ep_{p['id']}")] for p in ps]
    await update.message.reply_text(t(L(ctx), "ep_pick"), reply_markup=InlineKeyboardMarkup(kb))
    return S_EP_PICK


async def ep_pick(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    ctx.user_data["_epid"] = int(q.data[3:])
    kb = InlineKeyboardMarkup(
        [
            [InlineKeyboardButton("📋 Name", callback_data="ef_name")],
            [InlineKeyboardButton("🏢 Kunde", callback_data="ef_customer")],
            [InlineKeyboardButton("🏷 Kostenstelle", callback_data="ef_cost_center")],
        ]
    )
    await q.edit_message_text(t(L(ctx), "ep_what"), reply_markup=kb)
    return S_EP_FIELD


async def ep_field(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    ctx.user_data["_epf"] = q.data[3:]
    await q.edit_message_text(t(L(ctx), "ep_val"))
    return S_EP_VAL


async def ep_val(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ud = ctx.user_data
    with db() as c:
        c.execute(f"UPDATE projects SET {ud['_epf']}=? WHERE id=?", (update.message.text.strip(), ud["_epid"]))
    await update.message.reply_text(t(L(ctx), "ep_ok"))
    return ConversationHandler.END


async def cmd_delprojekt(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not ADM(ctx):
        await update.message.reply_text(t(L(ctx), "no_adm"))
        return ConversationHandler.END
    ps = get_projects()
    if not ps:
        await update.message.reply_text(t(L(ctx), "no_p"))
        return ConversationHandler.END
    kb = [[InlineKeyboardButton(f"❌ {p['name']}", callback_data=f"dp_{p['id']}")] for p in ps]
    await update.message.reply_text(t(L(ctx), "dp_pick"), reply_markup=InlineKeyboardMarkup(kb))
    return S_DP_PICK


async def dp_pick(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    pid = int(q.data[3:])
    p = get_project(pid)
    ctx.user_data["_dpid"] = pid
    ctx.user_data["_dpn"] = p["name"] if p else "?"
    kb = InlineKeyboardMarkup(
        [[InlineKeyboardButton(t(L(ctx), "yes"), callback_data="dy"), InlineKeyboardButton(t(L(ctx), "no"), callback_data="dn")]]
    )
    await q.edit_message_text(t(L(ctx), "dp_ask").format(n=ctx.user_data["_dpn"]), reply_markup=kb)
    return S_DP_CONFIRM


async def dp_confirm(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    if q.data == "dy":
        with db() as c:
            c.execute("UPDATE projects SET active=0 WHERE id=?", (ctx.user_data["_dpid"],))
        await q.edit_message_text(t(L(ctx), "dp_ok").format(n=ctx.user_data["_dpn"]))
    else:
        await q.edit_message_text(t(L(ctx), "cancel"))
    return ConversationHandler.END


async def cmd_rolle(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not ADM(ctx):
        await update.message.reply_text(t(L(ctx), "no_adm"))
        return ConversationHandler.END
    users = get_users()
    kb = [[InlineKeyboardButton(f"👤 {u['name']}", callback_data=f"ru_{u['id']}")] for u in users]
    await update.message.reply_text(t(L(ctx), "role_pick_user"), reply_markup=InlineKeyboardMarkup(kb))
    return S_ROLE_PICK_USER


async def role_pick_user(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    uid = int(q.data[3:])
    user = get_user_by_id(uid)
    if not user:
        await q.edit_message_text("Mitarbeiter nicht gefunden.")
        return ConversationHandler.END
    ctx.user_data["_ruid"] = uid
    ctx.user_data["_runame"] = user["name"]
    kb = InlineKeyboardMarkup(
        [
            [InlineKeyboardButton(f"👤 {t(L(ctx), 'role_employee')}", callback_data="rv_employee")],
            [InlineKeyboardButton(f"🦺 {t(L(ctx), 'role_foreman')}", callback_data="rv_foreman")],
            [InlineKeyboardButton(f"👑 {t(L(ctx), 'role_admin')}", callback_data="rv_admin")],
        ]
    )
    await q.edit_message_text(t(L(ctx), "role_pick_value").format(name=user["name"]), reply_markup=kb)
    return S_ROLE_PICK_VALUE


async def role_pick_value(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    role = q.data[3:]
    uid = ctx.user_data["_ruid"]
    with db() as c:
        c.execute("UPDATE users SET role=?, is_admin=? WHERE id=?", (role, 1 if role == "admin" else 0, uid))
    refreshed = get_user_by_id(uid)
    if refreshed:
        if uid == UID(ctx):
            sync_user(ctx, refreshed)
        try:
            await apply_user_menu(ctx.bot, refreshed["tid"], refreshed.get("role") or role)
        except Exception as exc:
            LOG.warning("Konnte Menü für %s nicht aktualisieren: %s", refreshed["name"], exc)
    await q.edit_message_text(t(L(ctx), "role_ok").format(name=ctx.user_data["_runame"], role=role_label(role, L(ctx))))
    return ConversationHandler.END


async def cmd_deluser(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not ADM(ctx):
        await update.message.reply_text(t(L(ctx), "no_adm"))
        return ConversationHandler.END
    users = [u for u in get_users() if u["id"] != UID(ctx)]
    if not users:
        await update.message.reply_text(t(L(ctx), "du_none"))
        return ConversationHandler.END
    kb = [[InlineKeyboardButton(f"{'👑' if u['role']=='admin' else ('🦺' if u['role']=='foreman' else '👤')} {u['name']}", callback_data=f"du_{u['id']}")] for u in users]
    await update.message.reply_text(t(L(ctx), "du_pick"), reply_markup=InlineKeyboardMarkup(kb))
    return S_DU_PICK


async def du_pick(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    uid = int(q.data[3:])
    u = get_user_by_id(uid)
    if not u:
        await q.edit_message_text(t(L(ctx), "du_none"))
        return ConversationHandler.END
    ctx.user_data["_duid"] = uid
    ctx.user_data["_dun"] = u["name"]
    kb = InlineKeyboardMarkup(
        [[InlineKeyboardButton(t(L(ctx), "yes"), callback_data="duy"), InlineKeyboardButton(t(L(ctx), "no"), callback_data="dun")]]
    )
    await q.edit_message_text(t(L(ctx), "du_ask").format(n=u["name"]), reply_markup=kb)
    return S_DU_CONFIRM


async def du_confirm(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    if q.data == "duy":
        with db() as c:
            c.execute("UPDATE users SET active=0 WHERE id=?", (ctx.user_data["_duid"],))
        await q.edit_message_text(t(L(ctx), "du_ok").format(n=ctx.user_data["_dun"]))
    else:
        await q.edit_message_text(t(L(ctx), "cancel"))
    return ConversationHandler.END


async def cmd_korrektur(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not UID(ctx):
        await update.message.reply_text("→ /start")
        return ConversationHandler.END
    current = current_user_row(ctx)
    since = (date.today() - timedelta(days=60)).isoformat()
    entries = get_entries(date_from=since, editable_by=current, limit=25)
    if not entries:
        await update.message.reply_text(t(L(ctx), "corr_none"))
        return ConversationHandler.END
    ctx.user_data["_corr_entries"] = {e["id"]: e for e in entries}
    kb = [[InlineKeyboardButton(entry_brief(e, L(ctx)), callback_data=f"ce_{e['id']}")] for e in entries]
    await update.message.reply_text(t(L(ctx), "corr_pick"), reply_markup=InlineKeyboardMarkup(kb))
    return S_CORR_PICK


async def corr_pick(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    entry_id = int(q.data[3:])
    e = ctx.user_data.get("_corr_entries", {}).get(entry_id) or get_entry(entry_id)
    if not e:
        await q.edit_message_text(t(L(ctx), "corr_none"))
        return ConversationHandler.END
    ctx.user_data["_corr_id"] = entry_id
    buttons = [[InlineKeyboardButton(f"🔁 {t(L(ctx), 'field_type')}", callback_data="cf_type")]]
    buttons.append([InlineKeyboardButton(f"📅 {t(L(ctx), 'field_date')}", callback_data="cf_edate")])
    if e["entry_type"] == "work":
        buttons.extend(
            [
                [InlineKeyboardButton(f"🕐 {t(L(ctx), 'field_start')}", callback_data="cf_stime")],
                [InlineKeyboardButton(f"🕑 {t(L(ctx), 'field_end')}", callback_data="cf_etime")],
                [InlineKeyboardButton(f"☕ {t(L(ctx), 'field_break')}", callback_data="cf_brk")],
            ]
        )
    else:
        buttons.append([InlineKeyboardButton(f"⏱ {t(L(ctx), 'field_hours')}", callback_data="cf_hours")])
    buttons.extend(
        [
            [InlineKeyboardButton(f"📝 {t(L(ctx), 'field_notes')}", callback_data="cf_notes")],
            [InlineKeyboardButton(f"🗑️ {t(L(ctx), 'field_delete')}", callback_data="cf_delete")],
        ]
    )
    await q.edit_message_text(entry_brief(e, L(ctx)) + "\n\n" + t(L(ctx), "corr_field"), reply_markup=InlineKeyboardMarkup(buttons))
    return S_CORR_FIELD


async def corr_field(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    field = q.data[3:]
    if field == "delete":
        kb = InlineKeyboardMarkup(
            [[InlineKeyboardButton(t(L(ctx), "yes"), callback_data="cd_yes"), InlineKeyboardButton(t(L(ctx), "no"), callback_data="cd_no")]]
        )
        await q.edit_message_text("⚠️ Eintrag wirklich löschen?", reply_markup=kb)
        return S_CORR_DELETE
    if field == "type":
        kb = InlineKeyboardMarkup(
            [
                [InlineKeyboardButton(f"🛠 {t(L(ctx), 'type_work')}", callback_data="ct_work")],
                [InlineKeyboardButton(f"🏖 {t(L(ctx), 'type_vacation')}", callback_data="ct_vacation")],
                [InlineKeyboardButton(f"🤒 {t(L(ctx), 'type_sick')}", callback_data="ct_sick")],
            ]
        )
        await q.edit_message_text(t(L(ctx), "corr_field"), reply_markup=kb)
        return S_CORR_TYPE
    ctx.user_data["_corr_field"] = field
    await q.edit_message_text(t(L(ctx), "corr_value"))
    return S_CORR_VALUE


async def corr_type(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    new_type = q.data[3:]
    entry_id = ctx.user_data["_corr_id"]
    e = get_entry(entry_id)
    if not e:
        await q.edit_message_text(t(L(ctx), "corr_none"))
        return ConversationHandler.END
    if new_type == "work" and e["project_id"] == get_absence_project_id():
        await q.edit_message_text("⚠️ Wechsel zu Arbeit bitte als neuen Arbeitseintrag anlegen, da ein Projekt benötigt wird.")
        return ConversationHandler.END
    payload = {"entry_type": new_type}
    if new_type in {"vacation", "sick"}:
        payload["project_id"] = get_absence_project_id()
        payload["hours"] = float(e["hours"])
    update_entry(entry_id, UID(ctx), **payload)
    await q.edit_message_text(t(L(ctx), "corr_ok"))
    return ConversationHandler.END


async def corr_value(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    field = ctx.user_data["_corr_field"]
    entry_id = ctx.user_data["_corr_id"]
    raw = update.message.text.strip()

    if field == "edate":
        value = parse_date(raw)
        if not value:
            await update.message.reply_text(t(L(ctx), "bad_d"))
            return S_CORR_VALUE
    elif field in {"stime", "etime"}:
        value = parse_time(raw)
        if not value:
            await update.message.reply_text(t(L(ctx), "bad_t"))
            return S_CORR_VALUE
    elif field == "brk":
        try:
            value = int(raw)
        except ValueError:
            await update.message.reply_text(t(L(ctx), "bad_b"))
            return S_CORR_VALUE
    elif field == "hours":
        value = parse_hours(raw)
        if value is None:
            await update.message.reply_text(t(L(ctx), "bad_h"))
            return S_CORR_VALUE
    else:
        value = "" if raw in {"/skip", t(L(ctx), "skip")} else raw

    update_entry(entry_id, UID(ctx), **{field: value})
    await update.message.reply_text(t(L(ctx), "corr_ok"))
    return ConversationHandler.END


async def corr_delete(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    if q.data == "cd_yes":
        delete_entry(ctx.user_data["_corr_id"])
        await q.edit_message_text(t(L(ctx), "corr_deleted"))
    else:
        await q.edit_message_text(t(L(ctx), "cancel"))
    return ConversationHandler.END


async def configure_bot_menu(app: Application) -> None:
    await app.bot.set_my_commands(commands_for_role("employee"))


def _require_template(path: Path, lang: str):
    if not XLSX:
        raise RuntimeError("openpyxl fehlt")
    if not path.exists():
        raise FileNotFoundError(t(lang, "template_missing").format(name=path.name))


def _fit_customer_rows(items: list[dict], capacity: int) -> list[dict]:
    if len(items) <= capacity:
        return items
    head = items[: capacity - 1]
    rest = items[capacity - 1 :]
    if not rest:
        return head
    combined = {
        "employee": " / ".join(i["employee"] for i in rest[:3]) + (" + weitere" if len(rest) > 3 else ""),
        "stime": min((i["stime"] for i in rest if i.get("stime")), default=""),
        "etime": max((i["etime"] for i in rest if i.get("etime")), default=""),
        "brk": sum(int(i.get("brk") or 0) for i in rest),
        "hours": round(sum(float(i.get("hours") or 0) for i in rest), 2),
    }
    head.append(combined)
    return head


def export_customer_week(project_id: int, kw: int, year: int) -> Path:
    _require_template(TEMPLATE_KUNDE_PATH, "de")
    project = get_project(project_id)
    if not project:
        raise RuntimeError("Projekt nicht gefunden")

    start = date.fromisocalendar(year, kw, 1)
    end = start + timedelta(days=6)
    entries = get_entries(project_id=project_id, entry_type="work", date_from=start.isoformat(), date_to=end.isoformat())
    if not entries:
        raise ValueError("Keine Daten")

    grouped: dict[date, list[dict]] = defaultdict(list)
    day_notes: dict[date, list[str]] = defaultdict(list)
    overflow_days: list[tuple[date, list[dict]]] = []

    for e in entries:
        d = datetime.strptime(e["edate"], "%Y-%m-%d").date()
        grouped[d].append(e)
        note = (e.get("notes") or "").strip()
        if note and note not in day_notes[d]:
            day_notes[d].append(note)

    wb = load_workbook(TEMPLATE_KUNDE_PATH)
    ws = wb[wb.sheetnames[0]]
    ws.title = f"KW{kw}_{year}"
    ws["C3"] = project["customer"]
    ws["C5"] = project["name"]
    ws["A6"] = year
    ws["A8"] = kw

    blocks = [
        (9, 13),
        (14, 18),
        (19, 23),
        (24, 28),
        (29, 33),
        (34, 38),
        (39, 42),
    ]

    for start_row, end_row in blocks:
        for rng in list(ws.merged_cells.ranges):
            if rng.min_col == 3 and rng.max_col == 3 and not (rng.max_row < start_row or rng.min_row > end_row):
                ws.unmerge_cells(str(rng))
        ws.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)

    for day_index, (start_row, end_row) in enumerate(blocks):
        current_day = start + timedelta(days=day_index)
        capacity = end_row - start_row + 1
        ws.cell(start_row, 3).value = None
        for row in range(start_row, end_row + 1):
            for col in (2, 4, 5, 6, 7):
                ws.cell(row, col).value = None
            ws.cell(row, 7).value = 0

        entries_day = sorted(
            grouped.get(current_day, []),
            key=lambda item: (item["employee"], item["stime"], item["etime"], item["id"]),
        )
        desc = " / ".join(day_notes.get(current_day, [])) or project["name"]
        if len(entries_day) > capacity:
            desc = f"{desc} (siehe Anhang)"
            overflow_days.append((current_day, entries_day[capacity:]))

        ws.cell(start_row, 3).value = desc

        for offset, item in enumerate(entries_day[:capacity]):
            row = start_row + offset
            ws.cell(row, 2).value = item["employee"]
            ws.cell(row, 4).value = excel_time_from_hhmm(item.get("stime"))
            ws.cell(row, 5).value = excel_time_from_break_minutes(int(item.get("brk") or 0))
            ws.cell(row, 6).value = excel_time_from_hhmm(item.get("etime"))
            ws.cell(row, 7).value = float(item.get("hours") or 0)

    if overflow_days:
        ws_extra = wb.create_sheet(f"Anhang_KW{kw}")
        ws_extra.append(["Datum", "Mitarbeiter", "Tätigkeit / Bemerkung", "Beginn", "Pause (Min)", "Ende", "Std."])
        for current_day, items in overflow_days:
            for item in items:
                ws_extra.append(
                    [
                        current_day.strftime("%d.%m.%Y"),
                        item["employee"],
                        (item.get("notes") or project["name"]).strip(),
                        item.get("stime"),
                        int(item.get("brk") or 0),
                        item.get("etime"),
                        float(item.get("hours") or 0),
                    ]
                )

    out = BASE_DIR / f"zeitflow_kunde_{project['customer'].replace(' ', '_')}_{kw}_{year}.xlsx"
    wb.save(out)
    return out


def export_internal_year(user_id: int, year: int) -> Path:
    _require_template(TEMPLATE_INTERN_PATH, "de")
    user = get_user_by_id(user_id)
    if not user:
        raise RuntimeError("Mitarbeiter nicht gefunden")

    start = date(year, 1, 1)
    end = date(year, 12, 31)
    entries = get_entries(user_id=user_id, date_from=start.isoformat(), date_to=end.isoformat())

    data_by_day: dict[date, list[dict]] = defaultdict(list)
    for e in entries:
        d = datetime.strptime(e["edate"], "%Y-%m-%d").date()
        data_by_day[d].append(e)

    wb = load_workbook(TEMPLATE_INTERN_PATH)
    year_sheet_name, month_sheets = normalise_internal_template_year(wb, year)
    ws_year = wb[year_sheet_name]
    ws_year["A2"] = f"Jahresliste {year}"
    ws_year["B3"] = user["name"]
    ws_year["B4"] = float(user.get("vacation_days") or 30)
    if ws_year["B5"].value is None:
        ws_year["B5"] = 0

    for month in range(1, 13):
        sheet_name = month_sheets.get(month)
        if not sheet_name:
            continue
        ws = wb[sheet_name]
        ws["A1"] = month
        ws["D1"] = year

        for row in range(4, 35):
            for col in range(2, 11):
                ws.cell(row, col).value = None

        current = date(year, month, 1)
        while current.month == month:
            row = 3 + current.day
            entries_day = sorted(data_by_day.get(current, []), key=lambda x: (x["entry_type"], x["stime"], x["project"]))
            work_entries = [e for e in entries_day if e["entry_type"] == "work"]
            vac_hours = round(sum(float(e["hours"] or 0) for e in entries_day if e["entry_type"] == "vacation"), 2)
            sick_hours = round(sum(float(e["hours"] or 0) for e in entries_day if e["entry_type"] == "sick"), 2)

            unique_projects = []
            for e in work_entries:
                short = project_short_name(e["project"])
                if short and short not in unique_projects:
                    unique_projects.append(short)
            labels = unique_projects[:]
            if vac_hours:
                labels.append("Urlaub")
            if sick_hours:
                labels.append("Krank")
            ws.cell(row, 2).value = "/".join(labels) if labels else None

            work_hours = [round(float(e["hours"] or 0), 2) for e in work_entries]
            for idx in range(3, 8):
                ws.cell(row, idx).value = None
            for idx, value in enumerate(work_hours[:4], start=3):
                ws.cell(row, idx).value = value
            overflow = work_hours[4:]
            if overflow:
                ws.cell(row, 7).value = round(sum(overflow), 2)

            total_work = round(sum(work_hours), 2)
            total = round(total_work + vac_hours + sick_hours, 2)
            ws.cell(row, 8).value = total if total else None

            daily_target = float(user.get("daily_target_hours") or 8.0)
            ws.cell(row, 9).value = round(vac_hours / daily_target, 2) if vac_hours else None
            ws.cell(row, 10).value = round(sick_hours / daily_target, 2) if sick_hours else None

            current += timedelta(days=1)

    out = BASE_DIR / f"zeitflow_intern_{user['name'].replace(' ', '_')}_{year}.xlsx"
    wb.save(out)
    return out


async def cmd_exportkunde(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not FOREMAN(ctx):
        await update.message.reply_text(t(L(ctx), "no_foreman"))
        return ConversationHandler.END
    customers = sorted({p["customer"] for p in get_projects() if p["customer"]})
    if not customers:
        await update.message.reply_text(t(L(ctx), "export_no"))
        return ConversationHandler.END
    ctx.user_data["_ek_customers"] = customers
    kb, page, max_page, total = build_customer_keyboard(customers, ctx.user_data.get("_ekc_page", 1))
    await update.message.reply_text(f"{t(L(ctx), 'pick_customer')} ({page}/{max_page}, {total} Kunden)", reply_markup=kb)
    ctx.user_data["_ekc_page"] = page
    ctx.user_data["_ekc_max_page"] = max_page
    return S_EK_CUSTOMER



async def ek_customer_page(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    try:
        page = int(q.data.split("_")[-1])
    except Exception:
        page = 1

    customers = ctx.user_data.get("_ek_customers", [])
    if not customers:
        await q.edit_message_text(t(L(ctx), "export_no"))
        return ConversationHandler.END

    kb, page, max_page, total = build_customer_keyboard(customers, page)
    await q.edit_message_text(f"{t(L(ctx), 'pick_customer')} ({page}/{max_page}, {total} Kunden)", reply_markup=kb)
    ctx.user_data["_ekc_page"] = page
    ctx.user_data["_ekc_max_page"] = max_page
    return S_EK_CUSTOMER

async def ek_customer(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    idx = int(q.data[3:])
    customers = ctx.user_data.get("_ek_customers", [])
    if idx >= len(customers):
        await q.edit_message_text(t(L(ctx), "export_no"))
        return ConversationHandler.END
    customer = customers[idx]
    ctx.user_data["_ek_customer"] = customer
    projects = [p for p in get_projects() if p["customer"] == customer]
    ctx.user_data["_ek_projects"] = projects
    ctx.user_data["_ekp_page"] = 1
    kb, page, max_page, total = build_export_project_keyboard(projects, ctx.user_data.get("_ekp_page", 1))
    await q.edit_message_text(f"{t(L(ctx), 'pick_project_export')} ({page}/{max_page}, {total} Projekte)", reply_markup=kb)
    ctx.user_data["_ekp_page"] = page
    ctx.user_data["_ekp_max_page"] = max_page
    return S_EK_PROJECT



async def ek_project_page(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    try:
        page = int(q.data.split("_")[-1])
    except Exception:
        page = 1

    projects = ctx.user_data.get("_ek_projects", [])
    if not projects:
        await q.edit_message_text(t(L(ctx), "export_no"))
        return ConversationHandler.END

    kb, page, max_page, total = build_export_project_keyboard(projects, page)
    await q.edit_message_text(f"{t(L(ctx), 'pick_project_export')} ({page}/{max_page}, {total} Projekte)", reply_markup=kb)
    ctx.user_data["_ekp_page"] = page
    ctx.user_data["_ekp_max_page"] = max_page
    return S_EK_PROJECT

async def ek_project(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    pid = int(q.data[4:])
    ctx.user_data["_ek_project_id"] = pid
    current_kw = date.today().isocalendar().week
    current_year = date.today().year
    await q.edit_message_text(f"{t(L(ctx), 'ask_kw')}\n\nAktuell: {current_kw}/{current_year}")
    return S_EK_WEEK


async def ek_week(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    parsed = parse_kw_year(update.message.text)
    if not parsed:
        await update.message.reply_text(t(L(ctx), "bad_kw"))
        return S_EK_WEEK
    kw, year = parsed
    try:
        path = export_customer_week(ctx.user_data["_ek_project_id"], kw, year)
    except FileNotFoundError as exc:
        await update.message.reply_text(str(exc))
        return ConversationHandler.END
    except ValueError:
        await update.message.reply_text(t(L(ctx), "export_no"))
        return ConversationHandler.END
    await update.message.reply_document(document=path.open("rb"), filename=path.name, caption=t(L(ctx), "exportkunde_done"))
    return ConversationHandler.END


async def cmd_exportintern(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not UID(ctx):
        await update.message.reply_text("→ /start")
        return ConversationHandler.END
    if FOREMAN(ctx):
        users = get_users()
        others = [u for u in users if u["id"] != UID(ctx)]
        ctx.user_data["_ei_users"] = others
        kb_self = [[InlineKeyboardButton(t(L(ctx), "self_export"), callback_data="iu_self")]]
        kb_users, page, max_page, total = build_export_user_keyboard(others, ctx.user_data.get("_ei_page", 1))
        kb = kb_self + kb_users
        await update.message.reply_text(f"{t(L(ctx), 'pick_export_user')} ({page}/{max_page}, {total} Mitarbeiter)", reply_markup=InlineKeyboardMarkup(kb))
        ctx.user_data["_ei_page"] = page
        ctx.user_data["_ei_max_page"] = max_page
        return S_EI_TARGET

    ctx.user_data["_ei_user_id"] = UID(ctx)
    await update.message.reply_text(f"{t(L(ctx), 'ask_export_year')}\n\nAktuell: {date.today().year}")
    return S_EI_YEAR



async def ei_user_page(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    try:
        page = int(q.data.split("_")[-1])
    except Exception:
        page = 1

    others = ctx.user_data.get("_ei_users", [])
    kb_self = [[InlineKeyboardButton(t(L(ctx), "self_export"), callback_data="iu_self")]]

    kb_users, page, max_page, total = build_export_user_keyboard(others, page)
    kb = kb_self + kb_users
    await q.edit_message_text(f"{t(L(ctx), 'pick_export_user')} ({page}/{max_page}, {total} Mitarbeiter)", reply_markup=InlineKeyboardMarkup(kb))
    ctx.user_data["_ei_page"] = page
    ctx.user_data["_ei_max_page"] = max_page
    return S_EI_TARGET

async def ei_target(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    payload = q.data[3:]
    if payload == "self":
        ctx.user_data["_ei_user_id"] = UID(ctx)
    else:
        ctx.user_data["_ei_user_id"] = int(payload)
    await q.edit_message_text(f"{t(L(ctx), 'ask_export_year')}\n\nAktuell: {date.today().year}")
    return S_EI_YEAR


async def ei_year(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if not raw.isdigit():
        await update.message.reply_text(t(L(ctx), "bad_year"))
        return S_EI_YEAR
    year = int(raw)
    if year < 2000 or year > 2100:
        await update.message.reply_text(t(L(ctx), "bad_year"))
        return S_EI_YEAR
    try:
        path = export_internal_year(ctx.user_data["_ei_user_id"], year)
    except FileNotFoundError as exc:
        await update.message.reply_text(str(exc))
        return ConversationHandler.END
    except ValueError:
        await update.message.reply_text(t(L(ctx), "export_no"))
        return ConversationHandler.END
    await update.message.reply_document(document=path.open("rb"), filename=path.name, caption=t(L(ctx), "exportintern_done"))
    return ConversationHandler.END


def main():
    token = os.environ.get("ZEITFLOW_BOT_TOKEN")
    if not token:
        LOG.error("ZEITFLOW_BOT_TOKEN fehlt!")
        return

    init_db()
    app = Application.builder().token(token).post_init(configure_bot_menu).build()
    fb = [CommandHandler("cancel", cancel)]

    app.add_handler(
        ConversationHandler(
            entry_points=[CallbackQueryHandler(reg_lang, pattern=r"^rl_")],
            states={S_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, reg_name)]},
            fallbacks=fb,
        )
    )

    app.add_handler(
        ConversationHandler(
            entry_points=[CommandHandler("zeit", cmd_zeit)],
            states={
                S_TARGET: [CallbackQueryHandler(z_target, pattern=r"^tu_")],
                S_TYPE: [CallbackQueryHandler(z_type, pattern=r"^ty_")],
                S_PROJ: [CallbackQueryHandler(proj_page_cb, pattern=r"^proj_page_"),
                CallbackQueryHandler(z_proj, pattern=r"^p_\d+$")],
                S_DATE: [
                    CallbackQueryHandler(z_date_btn, pattern=r"^dt$"),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, z_date_txt),
                ],
                S_START: [MessageHandler(filters.TEXT & ~filters.COMMAND, z_start)],
                S_END: [MessageHandler(filters.TEXT & ~filters.COMMAND, z_end)],
                S_BREAK: [MessageHandler(filters.TEXT & ~filters.COMMAND, z_break)],
                S_ABS_MODE: [CallbackQueryHandler(z_abs_mode, pattern=r"^am_")],
                S_ABS_HOURS: [MessageHandler(filters.TEXT & ~filters.COMMAND, z_abs_hours)],
                S_NOTES: [MessageHandler(filters.TEXT & ~filters.COMMAND, z_notes)],
                S_MORE: [CallbackQueryHandler(z_more, pattern=r"^m[yn]$")],
            },
            fallbacks=fb,
        )
    )

    app.add_handler(
        ConversationHandler(
            entry_points=[CommandHandler("addprojekt", cmd_addprojekt)],
            states={
                S_AP1: [MessageHandler(filters.TEXT & ~filters.COMMAND, ap1)],
                S_AP2: [MessageHandler(filters.TEXT & ~filters.COMMAND, ap2)],
                S_AP3: [MessageHandler(filters.TEXT & ~filters.COMMAND, ap3)],
            },
            fallbacks=fb,
        )
    )

    app.add_handler(
        ConversationHandler(
            entry_points=[CommandHandler("editprojekt", cmd_editprojekt)],
            states={
                S_EP_PICK: [CallbackQueryHandler(ep_pick, pattern=r"^ep_\d+$")],
                S_EP_FIELD: [CallbackQueryHandler(ep_field, pattern=r"^ef_")],
                S_EP_VAL: [MessageHandler(filters.TEXT & ~filters.COMMAND, ep_val)],
            },
            fallbacks=fb,
        )
    )

    app.add_handler(
        ConversationHandler(
            entry_points=[CommandHandler("delprojekt", cmd_delprojekt)],
            states={
                S_DP_PICK: [CallbackQueryHandler(dp_pick, pattern=r"^dp_\d+$")],
                S_DP_CONFIRM: [CallbackQueryHandler(dp_confirm, pattern=r"^d[yn]$")],
            },
            fallbacks=fb,
        )
    )

    app.add_handler(
        ConversationHandler(
            entry_points=[CommandHandler("rolle", cmd_rolle)],
            states={
                S_ROLE_PICK_USER: [CallbackQueryHandler(role_pick_user, pattern=r"^ru_\d+$")],
                S_ROLE_PICK_VALUE: [CallbackQueryHandler(role_pick_value, pattern=r"^rv_")],
            },
            fallbacks=fb,
        )
    )

    app.add_handler(
        ConversationHandler(
            entry_points=[CommandHandler("deluser", cmd_deluser)],
            states={
                S_DU_PICK: [CallbackQueryHandler(du_pick, pattern=r"^du_\d+$")],
                S_DU_CONFIRM: [CallbackQueryHandler(du_confirm, pattern=r"^du[yn]$")],
            },
            fallbacks=fb,
        )
    )

    app.add_handler(
        ConversationHandler(
            entry_points=[CommandHandler("korrektur", cmd_korrektur)],
            states={
                S_CORR_PICK: [CallbackQueryHandler(corr_pick, pattern=r"^ce_\d+$")],
                S_CORR_FIELD: [CallbackQueryHandler(corr_field, pattern=r"^cf_")],
                S_CORR_TYPE: [CallbackQueryHandler(corr_type, pattern=r"^ct_")],
                S_CORR_VALUE: [MessageHandler(filters.TEXT & ~filters.COMMAND, corr_value)],
                S_CORR_DELETE: [CallbackQueryHandler(corr_delete, pattern=r"^cd_")],
            },
            fallbacks=fb,
        )
    )

    app.add_handler(
        ConversationHandler(
            entry_points=[CommandHandler("exportkunde", cmd_exportkunde)],
            states={
                S_EK_CUSTOMER: [CallbackQueryHandler(ek_customer_page, pattern=r"^ecp_"),
                CallbackQueryHandler(ek_customer, pattern=r"^ec_\d+$")],
                S_EK_PROJECT: [CallbackQueryHandler(ek_project_page, pattern=r"^epj_page_"),
                CallbackQueryHandler(ek_project, pattern=r"^epj_\d+$")],
                S_EK_WEEK: [MessageHandler(filters.TEXT & ~filters.COMMAND, ek_week)],
            },
            fallbacks=fb,
        )
    )

    app.add_handler(
        ConversationHandler(
            entry_points=[CommandHandler("exportintern", cmd_exportintern)],
            states={
                S_EI_TARGET: [CallbackQueryHandler(ei_user_page, pattern=r"^iu_page_"),
                CallbackQueryHandler(ei_target, pattern=r"^iu_")],
                S_EI_YEAR: [MessageHandler(filters.TEXT & ~filters.COMMAND, ei_year)],
            },
            fallbacks=fb,
        )
    )

    for cmd, fn in [
        ("start", cmd_start),
        ("heute", cmd_heute),
        ("woche", cmd_woche),
        ("projekte", cmd_projekte),
        ("sprache", cmd_sprache),
        ("team", cmd_team),
        ("stats", cmd_stats),
        ("hilfe", cmd_hilfe),
        ("help", cmd_hilfe),
    ]:
        app.add_handler(CommandHandler(cmd, fn))

    app.add_handler(CallbackQueryHandler(sprache_cb, pattern=r"^sl_"))

    LOG.info("ZeitFlow Bot gestartet! DB: %s", DB_PATH)
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
